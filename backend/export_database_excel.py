import os
import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from sqlalchemy import inspect, text

# Import existing application database settings and engine
from app.config import settings
from app.database import engine

def get_postgresql_engine():
    """
    Ensures that exporter STRICTLY connects to the existing PostgreSQL database.
    DOES NOT fallback to SQLite or any other database.
    Raises RuntimeError if PostgreSQL connection fails.
    """
    db_url = settings.DATABASE_URL
    if not db_url or not db_url.startswith("postgresql"):
        raise RuntimeError(
            "[EXPORTER ERROR] PostgreSQL DATABASE_URL is not configured in settings. "
            f"Configured URL: '{db_url}'"
        )
    
    # Test connection to ensure PostgreSQL is reachable
    try:
        with engine.connect() as conn:
            conn.execute(text("SELECT 1;"))
        print("[POSTGRESQL EXPORTER] Successfully connected to PostgreSQL database.")
        return engine
    except Exception as e:
        raise RuntimeError(
            f"[EXPORTER CRITICAL ERROR] Failed to connect to PostgreSQL database using DATABASE_URL='{db_url}'.\n"
            f"Error details: {e}\n"
            "Exporter stopped: SQLite fallback is strictly disabled for database inspection exports."
        ) from e

def generate_database_excel_bytes() -> bytes:
    """
    READ-ONLY PostgreSQL Database Export to Excel (.xlsx)
    - Automatically identifies all existing PostgreSQL database tables.
    - Preserves column names, data types, NULLs, dates, and exact values.
    - Strictly READ-ONLY: DOES NOT insert, update, delete, or modify PostgreSQL schema/records.
    - Excludes database passwords/credentials from the output.
    """
    pg_engine = get_postgresql_engine()
    inspector = inspect(pg_engine)
    table_names = sorted(inspector.get_table_names())

    wb = openpyxl.Workbook()
    default_sheet = wb.active

    # Styling definitions for documentation spreadsheet
    header_fill = PatternFill(start_color="38A132", end_color="38A132", fill_type="solid")
    header_font = Font(name="Segoe UI", size=10, bold=True, color="FFFFFF")
    data_font = Font(name="Segoe UI", size=9, color="1A1410")
    empty_font = Font(name="Segoe UI", size=9, italic=True, color="7A6C5E")

    thin_border = Border(
        left=Side(style='thin', color='E2D7CB'),
        right=Side(style='thin', color='E2D7CB'),
        top=Side(style='thin', color='E2D7CB'),
        bottom=Side(style='thin', color='E2D7CB')
    )
    align_left = Alignment(horizontal='left', vertical='center')

    for index, table_name in enumerate(table_names):
        # Format sheet title (max 31 chars limit for Excel)
        sheet_title = table_name.replace("tbl_", "") if table_name.startswith("tbl_") else table_name
        sheet_title = sheet_title[:31]

        if index == 0 and default_sheet is not None:
            ws = default_sheet
            ws.title = sheet_title
        else:
            ws = wb.create_sheet(title=sheet_title)

        ws.views.sheetView[0].showGridLines = True

        # Retrieve column names from PostgreSQL schema
        columns_info = inspector.get_columns(table_name)
        column_names = [col['name'] for col in columns_info]

        # Write Header Row
        ws.append(column_names)
        header_row = ws[1]
        for cell in header_row:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = align_left
            cell.border = thin_border

        # READ-ONLY fetch of actual PostgreSQL records
        with pg_engine.connect() as conn:
            result = conn.execute(text(f'SELECT * FROM "{table_name}"'))
            rows = result.fetchall()

        if len(rows) == 0:
            empty_row = [None] * len(column_names)
            empty_row[0] = "No records currently stored in this table."
            ws.append(empty_row)
            row_idx = ws.max_row
            cell = ws.cell(row=row_idx, column=1)
            cell.font = empty_font
            cell.alignment = align_left
        else:
            for row in rows:
                formatted_row = []
                for val in row:
                    if val is None:
                        formatted_row.append(None)
                    elif isinstance(val, (int, float, bool)):
                        formatted_row.append(val)
                    else:
                        formatted_row.append(str(val))
                ws.append(formatted_row)
                row_idx = ws.max_row
                for col_idx in range(1, len(column_names) + 1):
                    c = ws.cell(row=row_idx, column=col_idx)
                    c.font = data_font
                    c.alignment = align_left
                    c.border = thin_border

        # Auto-adjust column widths
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                val_str = str(cell.value or '')
                if len(val_str) > max_len:
                    max_len = len(val_str)
            ws.column_dimensions[col_letter].width = min(max(max_len + 4, 12), 50)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()

def export_database_to_file(file_path: str = "RetailSphere_Database_Export.xlsx"):
    excel_bytes = generate_database_excel_bytes()
    with open(file_path, "wb") as f:
        f.write(excel_bytes)
    print(f"[EXPORT COMPLETE] PostgreSQL database export saved to: {os.path.abspath(file_path)}")
    
    # Also save to root if executed inside backend
    root_path = r"E:\Retail\RetailSphere_Database_Export.xlsx"
    with open(root_path, "wb") as f:
        f.write(excel_bytes)
    print(f"[EXPORT COMPLETE] PostgreSQL database export saved to: {root_path}")

if __name__ == "__main__":
    export_database_to_file()
